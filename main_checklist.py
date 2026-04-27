import json
import os
import re
import unicodedata
from difflib import get_close_matches
from openpyxl import load_workbook
from datetime import date


# ==============================
# 🔹 CONFIGURACIÓN (editar si es necesario)
# ==============================
JSON_PATH = "checklist/checkCV/check.json"
EXCEL_PATH = "checklist/ArquitectoExperto-ChecklistPlantilla.xlsx"
OUTPUT_PATH = "output.xlsx"
SHEET_NAME = None  # None = hoja activa


# ==============================
# 🔹 NORMALIZADOR DE TEXTO (más agresivo)
# ==============================
def normalizar(texto):
    if texto is None:
        return ""
    s = str(texto).lower().strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    # dejar solo letras, números y espacios
    s = re.sub(r"[^a-z0-9\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


# ==============================
# 🔹 APLANAR JSON (RECURSIVO)
# ==============================
def flatten_json(data):
    items = []

    # claves heurísticas para detectar campo concepto/resultado en nodos variados
    def detectar_claves(nodo):
        concepto_key = None
        resultado_key = None
        for k in nodo.keys():
            kn = normalizar(k)
            if ("concept" in kn) or ("pregunta" in kn) or ("item" in kn) or ("descripcion" in kn):
                concepto_key = k
            if ("result" in kn) or ("respuesta" in kn) or ("observacion" in kn) or ("valor" in kn) or ("estado" in kn):
                # preferir keys que no sean la misma que concepto
                if resultado_key is None or "result" in kn or "respuesta" in kn:
                    resultado_key = k
        return concepto_key, resultado_key

    def recorrer(nodo):
        if isinstance(nodo, list):
            for item in nodo:
                recorrer(item)

        elif isinstance(nodo, dict):
            ckey, rkey = detectar_claves(nodo)
            if ckey:
                concepto_val = nodo.get(ckey)
                resultado_val = None
                if rkey:
                    resultado_val = nodo.get(rkey)
                else:
                    # intentar buscar cualquier campo que parezca contener la respuesta
                    for k, v in nodo.items():
                        kn = normalizar(k)
                        if kn in ("valor", "respuesta", "resultado", "observacion", "estado"):
                            resultado_val = v
                            break

                items.append({"concepto": concepto_val, "resultado": resultado_val})

                # no recorrer más dentro de este nodo (ya extraímos lo que interesa)
            else:
                for value in nodo.values():
                    recorrer(value)

    recorrer(data)
    return items


def detectar_columnas(ws, claves_json):
    """Detecta automáticamente la columna del 'concepto' y la columna destino para 'resultado'.

    Estrategia:
    - Busca encabezados en la fila 1 con palabras clave.
    - Si no encuentra, prueba columnas 1..min(10, max_col) y elige la que más coincidencias tenga
      con las claves del JSON.
    - Resultado por defecto: concepto -> columna B (2), resultado -> columna E (5).
    """
    max_col = min(10, ws.max_column or 10)

    # 1) Buscar encabezados en fila 1
    header_map = {}
    for col in range(1, max_col + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            h = normalizar(v)
            header_map[h] = col

    # palabras clave comunes
    concepto_keys = ["concepto", "pregunta", "item", "descripcion", "descripciondelconcepto"]
    resultado_keys = ["resultado", "respuesta", "observacion", "observaciones", "estado", "veredicto"]

    for k in concepto_keys:
        if k in header_map:
            concepto_col = header_map[k]
            break
    else:
        concepto_col = None

    for k in resultado_keys:
        if k in header_map:
            resultado_col = header_map[k]
            break
    else:
        resultado_col = None

    # 2) Si no detectó por encabezado, buscar columna con más coincidencias con JSON
    if concepto_col is None:
        best_col = None
        best_score = 0
        for col in range(1, max_col + 1):
            score = 0
            for row in range(2, min(200, ws.max_row or 200) + 1):
                v = ws.cell(row=row, column=col).value
                if not v:
                    continue
                nv = normalizar(v)
                # coincidencia exacta
                if nv in claves_json:
                    score += 2
                    continue
                # coincidencia aproximada (útil cuando el JSON es pequeño)
                try:
                    from difflib import get_close_matches as _gcm
                    if claves_json and _gcm(nv, list(claves_json), n=1, cutoff=0.6):
                        score += 1
                except Exception:
                    pass
            if score > best_score:
                best_score = score
                best_col = col
        if best_score > 0:
            concepto_col = best_col

    # 3) Si aún no hay resultado_col, usar columna E (5) por defecto
    if resultado_col is None:
        resultado_col = 5

    # valores por defecto finales
    if concepto_col is None:
        concepto_col = 2

    return concepto_col, resultado_col


def get_top_left_merged(ws, row, col):
    """Devuelve la celda top-left (min_row,min_col) si (row,col) está dentro de un rango fusionado."""
    try:
        for cr in ws.merged_cells.ranges:
            if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
                return cr.min_row, cr.min_col
    except Exception:
        pass
    return row, col


def main():
    # 1. Cargar JSON
    if not os.path.exists(JSON_PATH):
        print(f"No se encuentra el archivo JSON: {JSON_PATH}")
        return

    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 2. Aplanar JSON
    items = flatten_json(data)
    print(f"Total items en JSON: {len(items)}")

    # 3. Crear mapa: concepto normalizado → resultado
    mapa = {}
    for item in items:
        clave = normalizar(item.get("concepto", ""))
        if clave:
            mapa[clave] = item.get("resultado", "")

    if not mapa:
        print("No se detectaron pares 'concepto'/'resultado' en el JSON.")
        return

    # 4. Abrir Excel
    if not os.path.exists(EXCEL_PATH):
        print(f"No se encuentra la plantilla Excel: {EXCEL_PATH}")
        return

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active if SHEET_NAME is None else wb[SHEET_NAME]

    # 5. Detectar columnas
    concepto_col, resultado_col = detectar_columnas(ws, set(mapa.keys()))
    print(f"Columna detectada - concepto: {concepto_col}, resultado: {resultado_col}")

    # 5.1 Escribir fecha del día junto a la etiqueta 'Fecha' si existe
    fecha_hoy = date.today().strftime("%d/%m/%Y")
    fecha_escrita = False
    # buscar en las primeras 30 filas para la etiqueta de fecha
    max_search_row = min(30, ws.max_row or 30)
    max_search_col = min(30, ws.max_column or 30)
    for r in range(1, max_search_row + 1):
        for c in range(1, max_search_col + 1):
            v = ws.cell(row=r, column=c).value
            if v and "fecha" in normalizar(v):
                try:
                    tr, tc = get_top_left_merged(ws, r, c + 1)
                    ws.cell(row=tr, column=tc).value = fecha_hoy
                    fecha_escrita = True
                except Exception:
                    pass
                break
        if fecha_escrita:
            break

    coincidencias = 0
    sin_coincidencia = []

    # 6. Recorrer filas y escribir resultados
    for r in range(2, (ws.max_row or 2) + 1):
        cell_val = ws.cell(row=r, column=concepto_col).value
        clave = normalizar(cell_val)
        if not clave:
            continue
        # preparar valor a escribir (siempre string)
        def escribir_resultado(fila, col, valor):
            # si la celda forma parte de un rango combinado, escribir en la celda superior izquierda
            for cr in ws.merged_cells.ranges:
                try:
                    if (fila, col) in cr:
                        fila = cr.min_row
                        col = cr.min_col
                        break
                except TypeError:
                    # algunos objetos pueden no soportar membership, seguir
                    pass
            try:
                ws.cell(row=fila, column=col).value = "" if valor is None else str(valor)
            except Exception:
                # como último recurso, convertir a string simple
                ws.cell(row=fila, column=col).value = str(valor)

        if clave in mapa:
            escribir_resultado(r, resultado_col, mapa[clave])
            coincidencias += 1
        else:
            # intentar coincidencia aproximada
            sugerencias = get_close_matches(clave, mapa.keys(), n=1, cutoff=0.8)
            if sugerencias:
                s = sugerencias[0]
                escribir_resultado(r, resultado_col, mapa[s])
                coincidencias += 1
            else:
                sin_coincidencia.append((r, cell_val, clave))

    # 7. Añadir hoja con no coincidencias para revisión
    if sin_coincidencia:
        if "SinCoincidencia" in wb.sheetnames:
            del wb["SinCoincidencia"]
        sheet_nc = wb.create_sheet("SinCoincidencia")
        sheet_nc.append(["Fila", "Concepto en Excel", "Concepto normalizado"])
        for fila, orig, norm in sin_coincidencia:
            sheet_nc.append([fila, orig, norm])

    # 8. Guardar resultado y verificar integridad básica reabriendo
    saved_path = None
    try:
        wb.save(OUTPUT_PATH)
        saved_path = OUTPUT_PATH
    except Exception as e:
        print(f"Error al guardar {OUTPUT_PATH}: {e}")
        # intentar guardar con sufijo de timestamp para evitar bloqueo por archivo abierto
        from datetime import datetime as _dt
        ts = _dt.now().strftime("%Y%m%d_%H%M%S")
        alt = OUTPUT_PATH.replace('.xlsx', f'_autosave_{ts}.xlsx')
        try:
            wb.save(alt)
            saved_path = alt
            print(f"Archivo guardado en ruta alternativa: {alt}")
        except Exception as e2:
            print(f"Fallo al guardar copia alternativa: {e2}")
            print("Asegúrate de que el archivo de salida no esté abierto en otra aplicación y vuelve a ejecutar el script.")

    # intentar reabrir para detectar corrupción inmediata si se guardó algo
    if saved_path:
        try:
            _ = load_workbook(saved_path)
        except Exception as e:
            alt2 = saved_path.replace('.xlsx', '_verificacion.xlsx')
            print(f"Advertencia: el archivo generado parece estar dañado al reabrir: {e}")
            print(f"Intentando guardar una copia alternativa: {alt2}")
            try:
                wb.save(alt2)
                saved_path = alt2
                print(f"Copia alternativa guardada en: {alt2}")
            except Exception as e3:
                print(f"Fallo al guardar copia alternativa: {e3}")

    print(f"Coincidencias encontradas: {coincidencias}")
    print(f"Filas sin coincidencia: {len(sin_coincidencia)} (ver hoja 'SinCoincidencia')")
    print(f"Archivo generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
